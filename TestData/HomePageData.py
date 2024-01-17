import openpyxl


class HomePageData:

    test_home_data = [{"firstName":"Rahul","email":"shetty@gmail.com","password":"123456", "gender":"Male"},{"firstName":"Anika","email":"anika@gmail.com","password":"56479", "gender":"Female"}]

    @staticmethod
    def getTestData(test_case_name):
        Diccionary = {}
        book = openpyxl.load_workbook("C:\\Users\\USER\\PycharmProjects\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Diccionary[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Diccionary]